#!/usr/bin/env python
# coding: utf-8

# In[7]:


# Webinspect Report Covert to Excel
# Copyright (c) 2023 JASON C < p477d343@gmail.com >

# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import tkinter as tk
from ttkbootstrap import Style
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
import openpyxl
import xlsxwriter
import os
import csv
import pandas as pd
import numpy as np
from os import path,listdir

style = Style(theme='flatly')
window = style.master


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("WebInspect 報告轉換器")
        self.root.resizable(False, False)
        
        # 建立標籤和文字輸入框
        self.path_label = tk.Label(root, text="路徑：", font=('helvetica', 12, 'bold'))
        self.path_label.grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.path_entry = tk.Entry(root)
        self.path_entry.grid(row=0, column=1, padx=10, pady=10)
        
        self.scandate_label = tk.Label(root, text="掃描年度(e.g. 2023.5)：", font=('helvetica', 12, 'bold'))
        self.scandate_label.grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.scandate_entry = tk.Entry(root)
        self.scandate_entry.grid(row=1, column=1, padx=10, pady=10)
        
        self.scantimes_label = tk.Label(root, text="本次為第幾次掃描(e.g. 二)：", font=('helvetica', 12, 'bold'))
        self.scantimes_label.grid(row=2, column=0, padx=10, pady=10, sticky="e")
        self.scantimes_entry = tk.Entry(root)
        self.scantimes_entry.grid(row=2, column=1, padx=10, pady=10)
        
        # 建立按鈕
        self.browse_button = tk.Button(root, text="瀏覽", command=self.browse_directory, width=10, font=('helvetica', 12))
        self.browse_button.grid(row=0, column=2, padx=10, pady=10)
        
        self.convert_button = tk.Button(root, text="轉換", command=self.convert, width=30, font=('helvetica', 12))
        self.convert_button.grid(row=3, column=1, padx=10, pady=10)
        
    def browse_directory(self):
        directory = filedialog.askdirectory()
        self.path_entry.delete(0, tk.END)
        self.path_entry.insert(0, directory)
        
    def convert(self):
        # 獲取使用者輸入值
        path = self.path_entry.get()
        scandate = self.scandate_entry.get()
        scantimes = self.scantimes_entry.get()
        
        # 執行轉換
        try:
            

            finalfilename =  path + "\\" + scandate + "網站弱掃_弱點匯整清冊_第" + scantimes + "次.xlsx"
            
            #!/usr/bin/env python
            # coding: utf-8

            # In[1]:


            site_name_list = []
            site_web_list = []
            web_link_list = []
            crisk_list = []
            hrisk_list = []
            mrisk_list = []


            firstsheet_dict = {"部門名稱":[],"網站名稱":[],"Website":[],"極高風險":[],"高風險":[],"中風險":[]};
            secondsheet_dict = {};
            thirddf = pd.DataFrame(data=secondsheet_dict)
            seconddf = pd.DataFrame(data=secondsheet_dict)
            files = listdir(path)
            list_sites= []
            for (root, dirs, file) in os.walk(path):
                for f in file:
                    if '.pdf' in f:
                        list_sites.append((str(root).split('\\'))[-1])

                for f in file:
                    if '.csv' in f:

                        file = str(root)+'\\'+str(f)
                        df = pd.read_csv(file,encoding='big5')

                        cols = df.columns.to_list()
                        myorder = [1, 4, 3, 2, 0, 5]
                        cols = [cols[i] for i in myorder]
                        df = df[cols]
                        col_name = df.columns.tolist()

                        col_name.insert(col_name.index('ReqMethod')+1,'new_cols')
                        df=df.reindex(columns=col_name)
                        df= df.replace(np.nan,' )')

                        col_name.insert(col_name.index('VulnCat')+1,'site_web_name')
                        df=df.reindex(columns=col_name)
                        df= df.replace(np.nan,root.split('\\')[-1])

                        col_name.insert(col_name.index('site_web_name')+1,'site')
                        df = df.reindex(columns=col_name)
                        df = df.replace(np.nan,root.split('\\')[-2])

                        df['VulnCat'] = df.VulnCat.str.cat(df.VulnName, sep=': ')
                        df['VulnCat'] = df.VulnCat.str.cat(df.VulnId.astype(str), sep=' ( ')
                        df['VulnCat'] = df.VulnCat.str.cat(df.new_cols, sep='')

                        df.drop(['VulnName','VulnId','ItemNumber','ReqMethod','new_cols'],axis=1,inplace=True)
                        df = df.set_index("Severity", drop=True)
                        df = df.drop("Low",errors='ignore')
                        df = df.reset_index()

                        cols = df.columns.to_list()

                        myorder = [3, 2, 0, 1]
                        cols = [cols[i] for i in myorder]
                        df = df[cols]

                        df = df.replace(to_replace=["_: "],value="", inplace=False,regex=True)

                        df = df.rename(columns = {'site':'部門名稱', 'site_web_name':'網站名稱', 'Severity':'風險等級', 'VulnCat':'弱點類別'})

                        site_name_list.append(root.split('\\')[-2])
                        site_web_list.append(root.split('\\')[-1])
                        web_link_list.append('https://')
                        crisk_list.append((df.風險等級 == "Critical").sum())
                        hrisk_list.append((df.風險等級 == "High").sum())
                        mrisk_list.append((df.風險等級 == "Medium").sum())

                        df = df.value_counts().reset_index(name='弱點數量')

                        thirddf = pd.concat([thirddf, df.value_counts(sort=False).to_frame()])
                        thirddf.drop(columns=thirddf.columns[-1], axis=1, inplace=True)

            new_list = {}
            for i in list_sites:
                if list_sites.count(i) >= 1:
                    new_list[i] = list_sites.count(i)

            firstdf = pd.DataFrame(new_list.items(), columns=['部門名稱', '網站數'])

            col_name = firstdf.columns.tolist()

            col_name.insert(col_name.index('網站數')+1,'極高風險')
            firstdf=firstdf.reindex(columns=col_name)

            col_name.insert(col_name.index('極高風險')+1,'高風險')
            firstdf=firstdf.reindex(columns=col_name)

            col_name.insert(col_name.index('高風險')+1,'中風險')
            firstdf = firstdf.reindex(columns=col_name)

            col_name.insert(col_name.index('中風險')+1,'備註')
            firstdf = firstdf.reindex(columns=col_name)


            firstsheet_dict.update({"部門名稱":site_name_list})
            firstsheet_dict.update({"網站名稱":site_web_list})
            firstsheet_dict.update({"Website":web_link_list})
            firstsheet_dict.update({"極高風險":crisk_list})
            firstsheet_dict.update({"高風險":hrisk_list})
            firstsheet_dict.update({"中風險":mrisk_list})

            thirddf.rename(index={'Critical':'極高'},inplace=True)
            thirddf.rename(index={'High':'高'},inplace=True)
            thirddf.rename(index={'Medium':'中'},inplace=True)

            seconddf = pd.DataFrame.from_dict(firstsheet_dict)

            fourth_dict = {'部門名稱': [],'弱點類別': [],'弱點合計': []}

            for i in range(0,len(thirddf.index)):
                fourth_dict["部門名稱"].append(thirddf.index[i][0])
                fourth_dict["弱點類別"].append(thirddf.index[i][3])
                fourth_dict["弱點合計"].append(thirddf.index[i][4])
            fourthdf = pd.DataFrame(fourth_dict)

            df2 = pd.concat([fourthdf, fourthdf.groupby(['弱點類別'],as_index=False)['弱點合計'].sum()]).sort_values('弱點類別')
            df2 = df2.dropna(axis=0)
            df2 = df2.values.tolist()

            with pd.ExcelWriter(finalfilename) as writer:
                firstdf.to_excel(writer, sheet_name='統計資料', index = False)
                seconddf.to_excel(writer, sheet_name='弱點彙總表', index = False)
                thirddf.to_excel(writer, sheet_name='網站弱點列表')

            wb = xlsxwriter.Workbook('十大高風險原始資料.xlsx')
            ws = wb.add_worksheet('十大高風險原始資料')

            headings=['部門名稱', '弱點類別','弱點合計']

            bold = wb.add_format({'bold': True})

            ws.write_row('A1', headings, bold)

            item = df2[0][1]
            rownum = 1
            startrow = 1
            count = 0
            for row in df2:
                count += 1
                if row[1] == item:
                    ws.set_row(rownum, None, None, {'level': 2, 'hidden': True})
                    ws.write_row(rownum, 0, row, bold)
                    rownum += 1
                else:
                    ws.set_row(rownum, None, None, {'level': 1})
                    ws.write(rownum, 1, item + ' 合計', bold)
                    cellno = 'C{}:C{}'.format(startrow, rownum)
                    ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')
                    item = df2[count-1][1]
                    rownum += 1
                    ws.set_row(rownum, None, None, {'level': 2, 'hidden': True})
                    ws.write_row(rownum, 0, row, bold)
                    rownum += 1
                    startrow = rownum
            else:
                ws.set_row(rownum, None, None, {'level': 1})
                ws.write(rownum, 1, item + " 合計")
                cellno = 'C{}:C{}'.format(startrow, rownum)
                ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')
                
            rownum += 1
            ws.write(rownum, 1, '弱點總合計', bold)
            cellno = 'C{}:C{}'.format(1, rownum)
            ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')
            wb.close()
            
            #SPLITLINE------------------------------------------------------------------------------------------------------
            
            wb = xlsxwriter.Workbook('網站掃描十大高風險.xlsx')
            ws = wb.add_worksheet('網站掃描十大高風險')

            headings=['RANK', '弱點類別','弱點合計']

            bold = wb.add_format({'bold': True})

            ws.write_row('A1', headings, bold)

            item = df2[0][1]
            rownum = 1
            startrow = 1
            count = 0
            for row in df2:
                count += 1
                if row[1] == item:
                    ws.set_row(rownum, None, None, {'level': 2, 'hidden': True})
                    ws.write_row(rownum, 0, row, bold)
                    rownum += 1
                else:
                    ws.set_row(rownum, None, None, {'level': 1})
                    ws.write(rownum, 1, item + ' 合計', bold)
                    cellno = 'C{}:C{}'.format(startrow, rownum)
                    ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')

                    item = df2[count-1][1]
                    rownum += 1
                    ws.set_row(rownum, None, None, {'level': 2, 'hidden': True})
                    ws.write_row(rownum, 0, row, bold)
                    rownum += 1
                    startrow = rownum
                    
            else:
                ws.set_row(rownum, None, None, {'level': 1})
                
                ws.write(rownum, 1, item + " 合計")
                cellno = 'C{}:C{}'.format(startrow, rownum)
                ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')
                
            rownum += 1
            ws.write(rownum, 1, '弱點總合計', bold)
            cellno = 'C{}:C{}'.format(1, rownum)
            ws.write(rownum, 2, '=SUBTOTAL(9,' + cellno + ')')
            wb.close()
            #SPLITLINE------------------------------------------------------------------------------------------------------
            

            import openpyxl
            
            xl1 = openpyxl.load_workbook(finalfilename)
            s = openpyxl.load_workbook('網站掃描十大高風險.xlsx').active
            s._parent = xl1
            xl1._add_sheet(s)

            s = openpyxl.load_workbook('十大高風險原始資料.xlsx').active
            s._parent = xl1
            xl1._add_sheet(s)

            xl1.save(finalfilename)
            

            from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
            from openpyxl import load_workbook

            wb = load_workbook(finalfilename)

            def set_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            ft= Font(name= "Microsoft JhengHei", size=12) 
            ft_bold = Font(name= "Microsoft JhengHei", size=12, bold=True)      

            #統計資料-----------------------------------------------------------------------------

            ws = wb[wb.sheetnames[0]]

            fixrownumberforsum = str(len(ws['B'])+1)
            fixrownumber = str(len(ws['B']))

            def set_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            set_border(ws, 'A2:F' + str(fixrownumberforsum))

            ws["B"+fixrownumberforsum] = "=SUM(B2:B"+fixrownumber+")"
            ws["C"+fixrownumberforsum] = "=SUM(C2:C"+fixrownumber+")"
            ws["D"+fixrownumberforsum] = "=SUM(D2:D"+fixrownumber+")"
            ws["E"+fixrownumberforsum] = "=SUM(E2:E"+fixrownumber+")"

            counternum = 2;
            for i in range(2,int(fixrownumberforsum)):
                webs_num = ws["B"+str(i)].value
                ws["C"+str(i)] = "=SUM(弱點彙總表!D" + str(counternum) + ":D" + str(counternum + int(ws["B" + str(i)].value)-1) + ")"
                ws["D"+str(i)] = "=SUM(弱點彙總表!E" + str(counternum) + ":E" + str(counternum + int(ws["B" + str(i)].value)-1) + ")"
                ws["E"+str(i)] = "=SUM(弱點彙總表!F" + str(counternum) + ":F" + str(counternum + int(ws["B" + str(i)].value)-1) + ")"
                counternum = counternum + webs_num

            ws["A"+fixrownumberforsum] = 'Total'


            for col in ws.columns:
                for cell in col:
                    cell.font = ft
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj

            ws.column_dimensions['A'].width = 25.0
            ws.column_dimensions['B'].width = 10.0
            ws.column_dimensions['C'].width = 10.0
            ws.column_dimensions['D'].width = 10.0
            ws.column_dimensions['E'].width = 10.0
            ws.column_dimensions['F'].width = 30.0

            ws.row_dimensions[1].height = 35.0
            ws.row_dimensions[int(fixrownumberforsum)].height = 25.0

            colorFill = PatternFill(start_color='E8E8D0',
                               end_color='E8E8D0',
                               fill_type='solid')
            darkredfill = PatternFill(start_color='AE0000',
                               end_color='AE0000',
                               fill_type='solid')
            redfill = PatternFill(start_color='EA0000',
                               end_color='EA0000',
                               fill_type='solid')
            yellowfill = PatternFill(start_color='F9F900',
                               end_color='F9F900',
                               fill_type='solid')

            for cell in ws["1:1"]:
                cell.fill = colorFill
                ws['C1'].fill = darkredfill
                ws['D1'].fill = redfill
                ws['E1'].fill = yellowfill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold

            for cell in ws[fixrownumberforsum+":"+fixrownumberforsum]:
                cell.fill = colorFill
                ws["C"+fixrownumberforsum].fill = colorFill
                ws["D"+fixrownumberforsum].fill = colorFill
                ws["E"+fixrownumberforsum].fill = colorFill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold    

            #統計資料-----------------------------------------------------------------------------

            #弱點彙總表-----------------------------------------------------------------------------

            ws = wb[wb.sheetnames[1]]

            fixrownumberforsum = str(len(ws['D'])+1)
            fixrownumber = str(len(ws['D']))

            set_border(ws, 'A2:F'+fixrownumber)
            set_border(ws, 'D'+fixrownumberforsum+':F'+fixrownumberforsum)

            ws["D"+fixrownumberforsum] = "=SUM(D2:D"+fixrownumber+")"
            ws["E"+fixrownumberforsum] = "=SUM(E2:E"+fixrownumber+")"
            ws["F"+fixrownumberforsum] = "=SUM(F2:F"+fixrownumber+")"

            for col in ws.columns:
                for cell in col:
                    cell.font = ft
                    alignment_obj = cell.alignment.copy(horizontal='left', vertical='center')
                    cell.alignment = alignment_obj

            ws.column_dimensions['A'].width = 20.0
            ws.column_dimensions['B'].width = 45.0
            ws.column_dimensions['C'].width = 80.0
            ws.column_dimensions['D'].width = 10.0
            ws.column_dimensions['E'].width = 10.0
            ws.column_dimensions['F'].width = 10.0
            ws.column_dimensions['D'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['E'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions['F'].alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[1].height = 35.0

            colorFill = PatternFill(start_color='E8E8D0',
                               end_color='E8E8D0',
                               fill_type='solid')
            darkredfill = PatternFill(start_color='AE0000',
                               end_color='AE0000',
                               fill_type='solid')
            redfill = PatternFill(start_color='EA0000',
                               end_color='EA0000',
                               fill_type='solid')
            yellowfill = PatternFill(start_color='F9F900',
                               end_color='F9F900',
                               fill_type='solid')

            for cell in ws["1:1"]:
                cell.fill = colorFill
                ws['D1'].fill = darkredfill
                ws['E1'].fill = redfill
                ws['F1'].fill = yellowfill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold

            for cell in ws[fixrownumberforsum+":"+fixrownumberforsum]:
                ws["D"+fixrownumberforsum].fill = colorFill
                ws["E"+fixrownumberforsum].fill = colorFill
                ws["F"+fixrownumberforsum].fill = colorFill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold    

            for cell in ws["D"]:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj 
            for cell in ws["E"]:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj  
            for cell in ws["F"]:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj  


            #弱點彙總表-----------------------------------------------------------------------------


            #網站弱點列表-----------------------------------------------------------------------------

            ws = wb[wb.sheetnames[2]]

            for col in ws.columns:
                for cell in col:
                    cell.font = ft
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj

            ws.column_dimensions['A'].width = 45.0
            ws.column_dimensions['B'].width = 45.0
            ws.column_dimensions['C'].width = 10.0
            ws.column_dimensions['D'].width = 80.0
            ws.column_dimensions['E'].width = 10.0

            ws.row_dimensions[1].height = 0.1
            ws.row_dimensions[2].height = 35.0

            for cell in ws["2:2"]:

                cell.fill = colorFill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold
            #網站弱點列表----------------------------------------------------------------------------------
            
            #網站掃描十大高風險-----------------------------------------------------------------------------
            ws = wb[wb.sheetnames[3]]
            
            fixrownumber = str(len(ws['C']))

            set_border(ws, 'A1:L'+fixrownumber)

            for col in ws.columns:
                for cell in col:
                    cell.font = ft_bold
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj
            ws['D1'] = "部門名稱"
            ws.merge_cells('D1:L1')
            ws.column_dimensions['A'].width = 20.0
            ws.column_dimensions['B'].width = 100.0
            ws.column_dimensions['C'].width = 10.0
            ws.column_dimensions['D'].width = 18.0
            ws.column_dimensions['E'].width = 18.0
            ws.column_dimensions['F'].width = 18.0
            ws.column_dimensions['G'].width = 18.0
            ws.column_dimensions['H'].width = 18.0
            ws.column_dimensions['I'].width = 18.0
            ws.column_dimensions['J'].width = 18.0
            ws.column_dimensions['K'].width = 18.0
            ws.column_dimensions['L'].width = 18.0

            ws.row_dimensions[1].height = 35.0

            for cell in ws["1:1"]:

                cell.fill = colorFill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold

            #網站掃描十大高風險-----------------------------------------------------------------------------

            #十大高風險原始資料-----------------------------------------------------------------------------
            ws = wb[wb.sheetnames[4]]

            fixrownumber = str(len(ws['C']))

            set_border(ws, 'A1:C'+fixrownumber)

            for col in ws.columns:
                for cell in col:
                    cell.font = ft_bold
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj

            ws.column_dimensions['A'].width = 20.0
            ws.column_dimensions['B'].width = 100.0
            ws.column_dimensions['C'].width = 10.0

            ws.row_dimensions[1].height = 35.0

            for cell in ws["1:1"]:

                cell.fill = colorFill
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
                cell.font = ft_bold

            #十大高風險原始資料-----------------------------------------------------------------------------
            os.remove('網站掃描十大高風險.xlsx')
            os.remove('十大高風險原始資料.xlsx')
            wb.save(finalfilename)
            
            
                    
            messagebox.showinfo("轉換完成", "轉換過程已完成。")
            
        except Exception as e:
            messagebox.showerror("錯誤", str(e))
        
root = tk.Tk()
app = App(root)
root.mainloop()


# In[ ]:




